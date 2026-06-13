#!/usr/bin/env python3
"""Convert an Excel salary spreadsheet to salary_data.json — CareerForge Epic 7 (T-061).

Usage:
  python3 tools/convert_salary_excel.py salary_data.xlsx
  python3 tools/convert_salary_excel.py salary_data.xlsx --source "Glassdoor 2025" --baseline 100

Requires:
  pip install openpyxl
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from typing import Dict, List, Optional, Tuple

try:
    import openpyxl
except ImportError:
    print(
        "Error: openpyxl is required.\n"
        "Install it with:  pip install openpyxl"
    )
    sys.exit(1)

# ---------------------------------------------------------------------------
# Column-name patterns (case-insensitive substring match)
# ---------------------------------------------------------------------------

COMPANY_PATTERNS = ["company", "virksomhed", "firma", "employer", "organization", "organisation"]
CITY_PATTERNS = ["city", "location", "by", "sted", "town", "region"]
COUNT_PATTERNS = ["count", "antal", "n ", "employees", "respondents"]
INDEX_PATTERNS = ["index", "indeks", "score", "ratio", "level"]

# Columns whose headers match these patterns (and are not company/city) are
# considered "data" columns eligible for pairing into categories.


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _is_numeric(value: object) -> bool:
    return isinstance(value, (int, float))


def _to_number(value: object) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", ".").strip())
    except (ValueError, TypeError):
        return None


def _matches_any(header: str, patterns: List[str]) -> bool:
    h = header.lower().strip()
    return any(p in h for p in patterns)


def _cell_str(cell: object) -> str:
    if cell is None:
        return ""
    return str(cell).strip()


# ---------------------------------------------------------------------------
# Header detection
# ---------------------------------------------------------------------------

def find_header_row(ws: "openpyxl.worksheet.worksheet.Worksheet") -> int:
    """Return 1-based row index of the header row (first row with >=2 non-empty string cells)."""
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), start=1):
        string_cells = [c for c in row if isinstance(c, str) and c.strip()]
        if len(string_cells) >= 2:
            return row_idx
    return 1  # Fallback


def detect_columns(
    headers: List[str],
) -> Tuple[int, Optional[int], List[Tuple[str, int, Optional[int]]]]:
    """Detect company col, optional city col, and category (name, index_col, count_col?) tuples.

    Returns:
        company_col: 0-based index
        city_col: 0-based index or None
        categories: list of (category_name, index_col_idx, count_col_idx_or_None)
    """
    company_col = 0  # Default to first column
    city_col: Optional[int] = None
    used: set = set()

    # Find company column
    for i, h in enumerate(headers):
        if _matches_any(h, COMPANY_PATTERNS):
            company_col = i
            used.add(i)
            break

    # Find city column
    for i, h in enumerate(headers):
        if i in used:
            continue
        if _matches_any(h, CITY_PATTERNS):
            city_col = i
            used.add(i)
            break

    # Remaining columns: pair count + index into categories
    # Strategy: scan left-to-right; if a column looks like a count, peek ahead for a
    # paired index column with a similar base name; otherwise treat it as a standalone index.
    remaining = [(i, h) for i, h in enumerate(headers) if i not in used and h.strip()]
    categories: List[Tuple[str, int, Optional[int]]] = []
    skip: set = set()

    for pos, (col_i, header_i) in enumerate(remaining):
        if col_i in skip:
            continue

        is_count = _matches_any(header_i, COUNT_PATTERNS)
        is_index = _matches_any(header_i, INDEX_PATTERNS)

        if is_count:
            # Look for a paired index column immediately to the right
            paired_idx = None
            if pos + 1 < len(remaining):
                next_col, next_header = remaining[pos + 1]
                if _matches_any(next_header, INDEX_PATTERNS):
                    paired_idx = next_col
                    skip.add(next_col)
            # Derive category name from the count header by stripping count-pattern words
            cat_name = header_i.strip()
            for p in COUNT_PATTERNS:
                cat_name = cat_name.lower().replace(p.strip(), "").strip()
            cat_name = cat_name.strip(" _-/") or "Category"
            categories.append((_title(cat_name), paired_idx if paired_idx is not None else col_i, col_i if paired_idx is None else paired_idx))
            # Rearrange so (name, index_col, count_col)
            if paired_idx is not None:
                categories[-1] = (_title(cat_name), paired_idx, col_i)
        elif is_index:
            # Standalone index column — use header as category name
            cat_name = header_i.strip()
            for p in INDEX_PATTERNS:
                cat_name = cat_name.lower().replace(p.strip(), "").strip()
            cat_name = _title(cat_name.strip(" _-/")) or _title(header_i.strip())
            categories.append((cat_name, col_i, None))
        else:
            # Generic numeric-looking column → treat as standalone index category
            categories.append((_title(header_i.strip()), col_i, None))

    return company_col, city_col, categories


def _title(s: str) -> str:
    """Title-case a string, preserving already-uppercase acronyms."""
    if not s:
        return s
    return s[0].upper() + s[1:]


# ---------------------------------------------------------------------------
# Row extraction
# ---------------------------------------------------------------------------

def extract_companies(
    ws: "openpyxl.worksheet.worksheet.Worksheet",
    header_row: int,
    company_col: int,
    city_col: Optional[int],
    categories: List[Tuple[str, int, Optional[int]]],
) -> List[dict]:
    companies = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if company_col >= len(row):
            continue
        company_name = _cell_str(row[company_col])
        if not company_name:
            continue

        city = _cell_str(row[city_col]) if city_col is not None and city_col < len(row) else None

        cats: Dict[str, dict] = {}
        for cat_name, idx_col, cnt_col in categories:
            index_val = _to_number(row[idx_col]) if idx_col < len(row) else None
            count_val = _to_number(row[cnt_col]) if cnt_col is not None and cnt_col < len(row) else None
            cats[cat_name] = {
                "count": int(count_val) if count_val is not None else None,
                "index": round(index_val, 2) if index_val is not None else None,
            }

        entry: dict = {"company": company_name}
        if city:
            entry["city"] = city
        entry["categories"] = cats
        companies.append(entry)

    return companies


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convert an Excel salary spreadsheet to salary_data.json.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("excel_file", help="Path to the Excel (.xlsx) file.")
    parser.add_argument(
        "--source",
        metavar="NAME",
        default=None,
        help="Data source label (default: filename without extension).",
    )
    parser.add_argument(
        "--baseline",
        type=int,
        default=100,
        metavar="N",
        help="Numeric baseline value (default: 100).",
    )
    parser.add_argument(
        "--baseline-label",
        default="index",
        metavar="LABEL",
        help="Label for the index column type (default: 'index').",
    )
    parser.add_argument(
        "--baseline-desc",
        default="Median = 100",
        metavar="DESC",
        help="Human-readable baseline description (default: 'Median = 100').",
    )
    parser.add_argument(
        "--output",
        default=None,
        metavar="PATH",
        help="Output JSON path (default: salary_data.json in the repo root).",
    )
    args = parser.parse_args()

    if not os.path.isfile(args.excel_file):
        print(f"Error: file not found: {args.excel_file}")
        sys.exit(1)

    source = args.source or os.path.splitext(os.path.basename(args.excel_file))[0]

    # Default output: repo root (parent of this script's directory)
    if args.output:
        output_path = args.output
    else:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        output_path = os.path.join(os.path.dirname(script_dir), "salary_data.json")

    print(f"Loading {args.excel_file} ...")
    wb = openpyxl.load_workbook(args.excel_file, data_only=True)
    ws = wb.active

    header_row = find_header_row(ws)
    raw_headers = [
        _cell_str(cell)
        for cell in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    ]
    print(f"  Header row: {header_row}  |  Columns: {[h for h in raw_headers if h]}")

    company_col, city_col, cat_cols = detect_columns(raw_headers)
    print(f"  Company col: {raw_headers[company_col]!r}")
    if city_col is not None:
        print(f"  City col:    {raw_headers[city_col]!r}")
    print(f"  Categories:  {[c[0] for c in cat_cols]}")

    companies = extract_companies(ws, header_row, company_col, city_col, cat_cols)
    print(f"  Rows extracted: {len(companies)}")

    data = {
        "metadata": {
            "source": source,
            "index_baseline": args.baseline,
            "index_label": args.baseline_label,
            "baseline_description": args.baseline_desc,
        },
        "companies": companies,
    }

    with open(output_path, "w", encoding="utf-8") as fh:
        json.dump(data, fh, indent=2, ensure_ascii=False)

    print(f"\nWrote {len(companies)} companies to {output_path}")


if __name__ == "__main__":
    main()
